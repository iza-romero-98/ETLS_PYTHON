import openpyxl
from openpyxl import workbook
import pyodbc
import json
from time import time
from datetime import date, datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import smtplib
import locale  
from os.path import basename
# from email import Encoders


def conectar_SQL():
    direccion_servidor = '74.208.235.157'
    # direccion_servidor = '217.160.15.235'
    # direccion_servidor = 'localhost'
    nombre_bd = 'InterERPv3P'
    nombre_usuario = 'sa'
    password = 'Root.inter2020!'

    try:
        conexion_SQL = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=' +
                                direccion_servidor+';DATABASE='+nombre_bd+';UID='+nombre_usuario+';PWD=' + password)
        print("OK! conexión exitosa")
    except Exception as e:
        # Atrapar error
        print("Ocurrió un error al conectar a SQL Server: ", e)
    
    return conexion_SQL

def informacion_beca(DP_CuentaDetalle, AL_ID, cursor):
    select_periodo_cpc = "select CPC_año, CPC_Periodo from CuentasPorCobrar inner join DetalleCuentaPorCobrar on "
    select_periodo_cpc += "CuentasPorCobrar.CPC_ID=DetalleCuentaPorCobrar.DCPC_CuentaId where DCPC_ID=" + str(DP_CuentaDetalle)  + ";"

    cursor.execute(select_periodo_cpc)
    periodo_cpc = cursor.fetchall()
    anio_ = periodo_cpc[0][0]
    periodo_ = periodo_cpc[0][1]
    

    select_beca = "select  beinscrip, beparcialidades, Becas_Nombre from becaAlumno inner join Becas on becaAlumno.Becas_ID=Becas.Becas_ID "
    select_beca += "where AL_ID =" + str(AL_ID).strip() + " and beanio=" +str(anio_).strip()+ " and beperiodo=" + str(periodo_).strip()+ ";"

    cursor.execute(select_beca)
    info_beca = cursor.fetchall()
    beinscrip_ = 0
    beparcialidades_ = 0
    bcdesc_ = ""
    if len(info_beca) > 0:
        beinscrip_ = info_beca[0][0]
        beparcialidades_ = info_beca[0][1]
        bcdesc_ = str(info_beca[0][2]).strip()

    return beinscrip_, beparcialidades_, bcdesc_

def descuento_cuenta(DP_CuentaDetalle, cursor, AP_Pagoid):
    select_desc_ = "select DCPC_ImporteTotal, DCPC_ID from DetalleCuentaPorCobrar where DCPC_ReferenciaCuentaDetalle=" + str(DP_CuentaDetalle)  + ";"
    cursor.execute(select_desc_)
    info_desc = cursor.fetchall()
    importe_desc_ = 0
    DCPC_ID = 0
    if len(info_desc) > 0:
        importe_desc_ = info_desc[0][0]    
        DCPC_ID = info_desc[0][1] #ID de cuenta de descuento

    flag_ = 0
    if DCPC_ID != 0:
        #Buscar si la cuenta está dentro del mismo pago:
        select_pago_ = "select DP_ID from DetallePago where DP_CuentaDetalle=" + str(DCPC_ID)  + "and DP_PagoID=" + str(AP_Pagoid) + ";"
        cursor.execute(select_pago_)
        info_desc_pago = cursor.fetchall()        
        if len(info_desc_pago) > 0:
            flag_= 1 #Bandera nos indica que es parte del mismo pago, entonces descontar a lo aplicado    
    
    return importe_desc_, flag_

def regimen(DF_Uso,cursor):

    if str(DF_Uso).strip() == "NULL" or str(DF_Uso).strip() == "None" or str(DF_Uso).strip()=="":
        UF_Descripcion_ = ""
    else:
        select_regimen = "select [UF_Descripcion] from [UsosFacturacion] where [UF_UsoID]='" + DF_Uso + "';"
        cursor.execute(select_regimen)
        UF_Descripcion_ =  cursor.fetchall()
        if len(UF_Descripcion_) > 0:
            return DF_Uso + " - " + UF_Descripcion_[0][0]
        else:
            return ""



def envia_reporte_facturacion(file_):
    # import necessary packages
    locale.setlocale(locale.LC_ALL, 'es-ES') 
    
    # create message object instance
    msg = MIMEMultipart()
    
    message = "Se adjunta el reporte de facturación del día " + str(date.today() + timedelta(days=-1)) + ". Este correo es generado automáticamente, no es necesario responder."
    
    with open(file_, "rb") as fil:
        attachedfile = MIMEApplication(fil.read(), _subtype = "xlsx")
        attachedfile.add_header('content-disposition', 'attachment', filename=basename(file_) )
        msg.attach(attachedfile)

    # setup the parameters of the message
    password = "desarrollom.2015"
    msg['From'] = "desarrollo@lainter.edu.mx"
    msg['To'] = "iy.romero@lainter.edu.mx"
    msg['Subject'] = "REPORTE DE FACTURACIÓN"
    
    # add in the message body
    msg.attach(MIMEText(message, 'html'))
    
    #create server
    server = smtplib.SMTP('smtp.gmail.com: 587')
    
    server.starttls()

    # Login Credentials for sending the mail
    server.login(msg['From'], password)
    
    
    # send the message via the server.
    server.sendmail(msg['From'], msg['To'], msg.as_string())    
    server.quit()

###################################################################3
####Main__ apartir de aquí corren los procesos 

conexion_SQL = conectar_SQL()

# Cursor SQL para obtener información de pagos 
with conexion_SQL.cursor() as cursor:
    fecha_ = date.today() + timedelta(days=-1)
    select_pagos = "select TOP(10) ap_pagoid, AP_FechaCreacion, AP_FechaContable, AP_AlumnoClave, AP_AlumnoID,AL_Nombre, AL_APP, AL_APM,"
    select_pagos += "AL_CURP, CarreraClave, Nivel_Nombre, AL_Factura, DP_CuentaDetalle,dcpc_descripcion, DCPC_ImporteTotal,"
    select_pagos += "DCPC_ImportePendiente,DP_ImporteAplicado,DCPC_ReferenciaCuentaDetalle,AP_AlumnoID, AP_AlumnoClave, AP_ImporteTotal,"
    select_pagos += "DF_RFC, DF_Nombre, DF_Direccion,DF_Email, DF_Uso,"
    select_pagos += "FP_Descripcion from AlumnoPagos "
    select_pagos += "inner join DetallePago on AlumnoPagos.AP_PagoID=DetallePago.DP_PagoID "
    select_pagos += "left join Alumno on AlumnoPagos.AP_AlumnoID=Alumno.AL_ID "
    select_pagos += "left join Carreras on Alumno.AL_Carrera = Carreras.IDCarrera "
    select_pagos += "left join Niveles on Carreras.Nivel_ID=Niveles.Nivel_ID "
    select_pagos += "left join DetalleCuentaPorCobrar on DetallePago.DP_CuentaDetalle=DetalleCuentaPorCobrar.DCPC_ID "
    select_pagos += "left join DatosFacturacion on AlumnoPagos.AP_AlumnoID=DatosFacturacion.AL_ID "
    select_pagos += "left join FormaPago on AlumnoPagos.AP_FormaPagoID=FormaPago.FP_ID "
    select_pagos += "where AP_ImporteTotal>0 and DCPC_ReferenciaCuentaDetalle=0 and AP_AlumnoID>0 and "
    select_pagos += "AP_FechaCreacion='2022-07-07' order by AP_PagoID;"

    # select_pagos += "(AP_FechaCreacion>='2022-06-22' and  AP_FechaCreacion<='2022-07-15') order by AP_PagoID;"

    # select_pagos += "AP_FechaCreacion='" + str(fecha_.year).strip() + "-" + str(fecha_.month).strip() + "-" + str(fecha_.day).strip() + "' order by AP_PagoID;"

    cursor.execute(select_pagos)
    informacion_pagos = cursor.fetchall()

    if len(informacion_pagos) > 0:
        #Archivo de facturación:
        wb = openpyxl.load_workbook('FormatoFacturacion.xlsx')
        # Active worksheet
        ws = wb.active
        ws['H2'] = datetime.now()
        fila_ = 4
        AP_AlumnoClave_ = ""
        for item in informacion_pagos:
            
            # print(item[3])
            AP_Pagoid = item[0]

            # Información de beca:
            if AP_AlumnoClave_ != str(item[3]).strip():
                beinscrip_, beparcialidades_, bcdesc_ = informacion_beca(item[12],item[18],cursor)
            #---------------------------------------------------------------------------------------------

            # Información descuento alumno
            importe_desc_, flag_ = descuento_cuenta(item[12],cursor,AP_Pagoid)
            print(item[16])

            if int(item[16]) > 0:
                importeaplicado_ = int(item[16]) 
            else: 
                importeaplicado_ = int(item[16])*-1

            if flag_ == 1: #Están dentro del mismo pago y es un descuento de beca
                importeaplicado_ = importeaplicado_ + importe_desc_
            #---------------------------------------------------------------------------------------------

            # Información importe de cuenta
            if importe_desc_ != 0:
                DCPC_ImporteTotal_ = int(item[14]) + importe_desc_
            else:
                DCPC_ImporteTotal_ = int(item[14]) + importe_desc_
            #---------------------------------------------------------------------------------------------


            AP_AlumnoClave_ = str(item[3]).strip()

            #Información  régimen
            UF_Descripcion_ = regimen(str(item[25]).strip(),cursor)

            # Escribir en excel:     
            # Información de pagos
            ws['B' + str(fila_)] = int(item[0])
            ws['C' + str(fila_)] = str(item[2]).strip()
            ws['D' + str(fila_)] = "SI" if item[11] == 1 else "NO"
            ws['E' + str(fila_)] = str(item[21]).strip() if (str(item[21]).strip() != "NULL" and str(item[21]).strip() != "None") else "" 
            ws['F' + str(fila_)] = str(item[22]).strip() if (str(item[22]).strip() != "NULL" and str(item[22]).strip() != "None") else "" 
            ws['G' + str(fila_)] = UF_Descripcion_
            ws['H' + str(fila_)] = str(item[23]).strip() if (str(item[23]).strip() != "NULL" and str(item[23]).strip() != "None") else ""

            ws['I' + str(fila_)] = str(item[26]).strip()
            ws['J' + str(fila_)] = "PUE" if DCPC_ImporteTotal_ == importeaplicado_ else "PPD"
            ws['K' + str(fila_)] = str(item[13]).strip()
            ws['L' + str(fila_)] = str(item[5]).strip() + " " + str(item[6]).strip() + " " + str(item[7]).strip()
            ws['M' + str(fila_)] = str(item[10]).strip()
            ws['N' + str(fila_)] = str(item[3]).strip()
            ws['O' + str(fila_)] = str(item[8]).strip()
            ws['P' + str(fila_)] = importeaplicado_ #int(item[16])*-1
            ws['Q' + str(fila_)] = int(item[14])
            ws['R' + str(fila_)] = importeaplicado_ #int(item[16])*-1
            ws['S' + str(fila_)] = int(item[15])
            ws['T' + str(fila_)] = str(beparcialidades_).strip()+'%' if beparcialidades_ > 0 else ""
            ws['U' + str(fila_)] = int(importe_desc_*-1) if (importe_desc_*-1) > 0 else ""
            ws['V' + str(fila_)] = str(bcdesc_).strip()
            ws['W' + str(fila_)] = str(item[24]).strip() if (str(item[24]).strip() != "NULL" and str(item[24]).strip() != "None") else ""

            fila_ += 1

        # Save the file
        wb.save("FormatoFacturacion_" + str(date.today()).strip()+ ".xlsx")
        # envia_reporte_facturacion("FormatoFacturacion_" + str(date.today()).strip()+ ".xlsx")
    
cursor.close()
conexion_SQL.close()